# -*- coding: utf-8 -*-
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import numpy as np
import pathlib
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Protection
import cv2 as cv
import glob,os,codecs
from mylib import utilxls


def makeheader(ws_dst_sht):
    ws_dst_sht.cell(1,1).value = "No."
    ws_dst_sht.cell(1,2).value = "Graph"
    ws_dst_sht.cell(1,3).value = "data1"
    ws_dst_sht.cell(1,4).value = "data2"
    ws_dst_sht.cell(1,5).value = "data3"
    ws_dst_sht.cell(1,6).value = "data4"
    ws_dst_sht.cell(1,7).value = "data5"
    return

def makepicexcel(imgdir, ws_sht, ws_dst_sht) :
    # idwpath = pathlib.Path("./" + imgdir)
    # print(idwpath)

    #adjust sheet Height
    # Excel cellサイズ：https://mainichi.doda.jp/article/2019/10/31/1735.html
    # セルの高さ設定「ポイント」は長さの単位で、１ポイント＝1/72インチ≒1.3ピクセル
    ws_dst_sht.column_dimensions['B'].width = 30     # set cell width 
    for idx in range(1, ws_sht.max_row):
        ws_dst_sht.row_dimensions[idx+1].height = 93  # set cell height 
    
    # ディレクトリ配下のテキストファイル一覧を表示（拡張子指定）
    # glob参考 https://www.fenet.jp/dotnet/column/language/7225/
    # Excel 画像挿入参考 https://qiita.com/quryu/items/7e281dcb11b0e3db3a99
    sortedfile = glob.glob(imgdir + "/*.png")
    sortedfile = sorted(sortedfile, key=utilxls.natural_keys)

    for idx, imgfname in enumerate(sortedfile):
        print(idx, imgfname)
        img = openpyxl.drawing.image.Image(imgfname)
        img.width = 240
        img.height = 120
        ws_dst_sht.add_image(img,"B%d" % (idx+2))

        ws_dst_sht.cell(idx+2, 1).value = ws_sht.cell(idx+2, 1).value
        ws_dst_sht.cell(idx+2, 3).value = ws_sht.cell(idx+2, 2).value
        ws_dst_sht.cell(idx+2, 4).value = ws_sht.cell(idx+2, 3).value
        ws_dst_sht.cell(idx+2, 5).value = ws_sht.cell(idx+2, 4).value
        ws_dst_sht.cell(idx+2, 6).value = ws_sht.cell(idx+2, 5).value
        ws_dst_sht.cell(idx+2, 7).value = ws_sht.cell(idx+2, 6).value

    return
    
def main():
    x_data = np.arange(1,10) # x軸のメモリ用
    # print(x_data)
    # ※np.loadtxtでは、ヘッダに文字があるとエラーとなってしまう。
    # df_raw = np.loadtxt("test3-2.csv", delimiter = ",")

    # 参考： https://gammasoft.jp/support/how-to-use-openpyxl-for-excel-file/
    file_path = './test3-3.xlsx'
    wb_mst = openpyxl.load_workbook(file_path)  # 
    ws_sht = wb_mst['test3-3']  # シート名を指定　
    # シートをインデックス番号で指定する場合
    # ws = wb.worksheets[0]     # インデックス番号は、0 から開始

    sht_maxRow = ws_sht.max_row         #最大行取得
    print("sht_maxRow", sht_maxRow)

    img_fname = []

    # for idx in range(sht_maxRow):
    for row in ws_sht.iter_rows(min_row=2) :
        ylist = []
        # cell アクセス
        # C1 = ws.cell("A1")            # 「A1」Cellのアクセス
        # c1 = ws.cell(row=1, column=1) # 「A1」を行列の番号で取得（注意：１から始まる）
        # c1 = ws.cell(1, 1)            #  キーワード（row=, column=）は省略も可能
        #「A1:C3」の範囲
        # rng1 = ws["A1:C3"]

        for cell in row :
            # ylist = ws_sht.cell(idx+1, 2).value
            ylist.append(cell.value)
        
        print(":", ylist)

        plt.figure(figsize=(16,10), dpi=80)
        # plt.plot(x_data[0:5], ylist_data[1:6])
      
        plt.bar(x_data[0:5], ylist[1:6], color=cm.Set1.colors[2], width=0.2, label="data/value")

        plt.title("データグラフ", fontsize=22, fontname="MS Gothic" )
        plt.xlabel('# Xvalue')
        plt.ylabel('# Yvalue')
        # plt.xlim(0, 6)
        plt.ylim(1, 50001) # 50000 -> 50001 Max目盛りに「50000」を表示したいため
        plt.xticks(np.arange(0, 6, step=1)) 
        plt.yscale('linear')    # linear     log
        plt.yticks(np.arange(0, 50001, step=10000)) # 50000 -> 50001 Max目盛りに「50000」を表示したいため

        dbgdisp = False # True # False
        if dbgdisp :
            plt.show()
        else:
            # plt.show()
            
            savefname = "./img/img" + str(ylist[0]) + ".png"
            # print("save File name:" + savefname + "\n")
            plt.savefig(savefname)
            img_fname.append(savefname)

    movname = "temp.mp4"
    utilxls.makemovie(img_fname, movname)

    imgdir = "img/"
    xlsfname = "./test3-3a.xlsx"    # 新規作成のExcelファイル名
    wb_dst = openpyxl.Workbook()
    ws_dst_sht = wb_dst.worksheets[0]  # シートを指定　

    makeheader(ws_dst_sht)      # Excelのヘッダ行追加
    ws_dst_sht = makepicexcel(imgdir, ws_sht, ws_dst_sht)
    wb_dst.save(xlsfname)

    print("end")

if __name__ == '__main__':
    main()
