# -*- coding: utf-8 -*-
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import numpy as np
import pathlib
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Protection
import cv2 as cv
import glob,os,codecs

def makemovie(img_fname):
    ### Movie build
    size = (0,0)
    frame_rate = 20

    #height = 800
    #width = 1280
    height = 400
    width = 640
    size = (width, height)
    # movname = './' + file_name + '.mp4'
    codec = cv.VideoWriter_fourcc('m','p','4','v')
    video = cv.VideoWriter('temp.mp4', codec, frame_rate, size)

    for fname in img_fname:
        print(fname)
        img = cv.imread(fname)
        img = cv.resize(img, dsize=(640, 400))
        video.write(img)

    video.release()
    
def main():
    x_data = np.arange(1,10) # x軸のメモリ用
    # print(x_data)

    # ※np.loadtxtでは、ヘッダに文字があるとエラーとなってしまう。
    # df_raw = np.loadtxt("test3-2.csv", delimiter = ",")

    # 参考： https://gammasoft.jp/support/how-to-use-openpyxl-for-excel-file/
    file_path = './test3-2.xlsx'
    wb_mst = openpyxl.load_workbook(file_path)  # 
    ws_sht = wb_mst['test3-2']  # シート名を指定　
    # シートをインデックス番号で指定する場合
    # ws = wb.worksheets[0]     # インデックス番号は、0 から開始

    sht_maxRow = ws_sht.max_row         #最大行取得
    print("sht_maxRow", sht_maxRow)

    img_fname = []

    # for idx in range(sht_maxRow):
    for row in ws_sht.iter_rows(min_row=2) :
        ylist = []
        # cell アクセス
        # C1 = ws.cell("A1")            # 「A1」Cellのアクセス
        # c1 = ws.cell(row=1, column=1) # 「A1」を行列の番号で取得（注意：１から始まる）
        # c1 = ws.cell(1, 1)            #  キーワード（row=, column=）は省略も可能
        #「A1:C3」の範囲
        # rng1 = ws["A1:C3"]

        for cell in row :
            # ylist = ws_sht.cell(idx+1, 2).value
            ylist.append(cell.value)
        
        print(":", ylist)

        plt.figure(figsize=(16,10), dpi=80)
        # plt.plot(x_data[0:5], ylist_data[1:6])
      
        plt.bar(x_data[0:5], ylist[1:6], color=cm.Set1.colors[2], width=0.2, label="data/value")

        plt.title("データグラフ", fontsize=22, fontname="MS Gothic" )
        plt.xlabel('# Xvalue')
        plt.ylabel('# Yvalue')
        # plt.xlim(0, 6)
        plt.ylim(1, 50001) # 50000 -> 50001 Max目盛りに「50000」を表示したいため
        plt.xticks(np.arange(0, 6, step=1)) 
        plt.yscale('linear')    # linear     log
        plt.yticks(np.arange(0, 50001, step=10000)) # 50000 -> 50001 Max目盛りに「50000」を表示したいため

        dbgdisp = False # True # False
        if dbgdisp :
            plt.show()
        else:
            # plt.show()
            savefname = "./img/img" + str(ylist[0]) + ".png"
            # print("save File name:" + savefname + "\n")
            plt.savefig(savefname)
            img_fname.append(savefname)

    makemovie(img_fname)
    print("end")

if __name__ == '__main__':
    main()
